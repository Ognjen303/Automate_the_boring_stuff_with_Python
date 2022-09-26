# USAGE: Run in command line--> python3 ch-13e.py <excel file name>

#! python3

import openpyxl, sys
from pathlib import Path

# Path that contians the desired .xlsx files
# change to whatever is suitable for you
p = Path(r'C:\automate the boring stuff fajlovi\automate_online-materials')

xlsxFile = sys.argv[-1]

wb = openpyxl.load_workbook(p / xlsxFile)
sheet = wb.active

maxRow = sheet.max_row
maxCol = sheet.max_column

for colNum in range(1, sheet.max_column + 1):
    fileName = 'excelToTextCol' + str(colNum) + '.txt'
    file = open(fileName, 'w')
    for rowNum in range(1, sheet.max_row + 1):
        file.write(sheet.cell(rowNum, colNum),value)
    file.close
