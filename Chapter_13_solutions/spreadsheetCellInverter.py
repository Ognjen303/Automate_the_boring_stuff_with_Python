#! python3

# spreadsheetCellInverter.py - returns the transpose of spreadsheet
# input type win + R: spreadsheetCellInverter.py

import openpyxl, sys
import numpy as np
from pathlib import Path

# path of imput .xlsx file, change to whatever is suitable for you
xlsxFilePath = Path(r'C:\automate the boring stuff fajlovi\automate_online-materials')

xlsxFile = sys.argv[-1]

wb = openpyxl.load_workbook(xlsxFilePath / xlsxFile)
sheet = wb.active
sheet.title = 'Before'

maxRow = sheet.max_row
maxCol = sheet.max_column

# create a new sheet to write to output
afterSheet = wb.create_sheet(index = 1, title = 'After')



for rowNum in range(1, maxCol + 1):
    for colNum in range(1, maxRow + 1):

        afterSheet.cell(rowNum, colNum).value = sheet.cell(colNum, rowNum).value

print('Saving sheet')
wb.save(xlsxFilePath / f'copy_of_{xlsxFile}')
wb.close()
print('Done')
