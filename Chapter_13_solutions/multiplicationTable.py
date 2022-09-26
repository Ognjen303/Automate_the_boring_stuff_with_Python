#! python3
# multiplicationTable.py - Takes in number N > 0 and creates a NxN multiplication table in Excel

import sys, openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

N = int(sys.argv[-1]) # last entry in sys.argv is N

wb = openpyxl.Workbook()
sheet = wb.active
boldFont = Font(bold = True)


for i in range(1, N + 1):

    # enter values of first 'header' row and column
    sheet.cell(row = 1, column = i + 1).value = i
    sheet.cell(row = 1, column = i + 1).font = boldFont

    sheet.cell(row = i + 1, column = 1).value = i
    sheet.cell(row = i + 1, column = 1).font = boldFont



    for j in range(1, i + 1):

        cell1 = 'A' + str(i + 1)
        cell2 = get_column_letter(j + 1) + '1'
        sheet.cell(row = i + 1, column = j + 1).value = f'=PRODUCT({cell1}, {cell2})'
        sheet.cell(row = j + 1, column = i + 1).value = f'=PRODUCT({cell1}, {cell2})'


wb.save('multiplicationTable.xlsx')
print('Saved table')
