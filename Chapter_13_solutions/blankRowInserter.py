#! python3

# blankRowInserter.py

# this is how the input looks like: blankRowInserter.py 3 2 myProduce.xlsx

import openpyxl, sys, os
from pathlib import Path

original_path = os.getcwd()

# path of imput .xlsx file, change to whatever is suitable for you
xlsxFilePath = Path(r'C:\automate the boring stuff fajlovi\automate_online-materials')

try:
	N = int(sys.argv[1])
	M = int(sys.argv[2])
	xlsxFile = sys.argv[3]

except ValueError as e:
	print('Both m and n should be integers')
	sys.exit(-2)


os.chdir(xlsxFilePath)
if not os.path.exists(xlsxFile):
	print('The specified file does not exist')
	sys.exit(-3)


xlsxFile = sys.argv[-1]


wb = openpyxl.load_workbook(xlsxFilePath / xlsxFile)
sheet = wb.active

outputWb = openpyxl.Workbook()
outputSheet = outputWb.active


for rowNum in range(1, sheet.max_row + 1):
    for colNum in range(1, sheet.max_column + 1):

        if rowNum < N:
            outputSheet.cell(rowNum, colNum).value = sheet.cell(rowNum, colNum).value

        else:
            outputSheet.cell(rowNum + M, colNum).value = sheet.cell(rowNum, colNum).value

os.chdir(original_path + original_path)
print('Saving table')
outputWb.save(f'{xlsxFile}_withBlankRows.xlsx')
input('Press ENTER to exit')
